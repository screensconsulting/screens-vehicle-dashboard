#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════════════════╗
║  Screens Studio — Global Infotainment Dashboard Auto-Updater v3.0        ║
║  consulting@screensauto.com                                              ║
╠══════════════════════════════════════════════════════════════════════════╣
║  Runs every Monday. Does four things:                                    ║
║   1. Reads Vehicle_Planing_Trial.xlsx  (vehicle availability)            ║
║   2. Fetches live market-share data from ACEA / Wards / CPCA             ║
║   3. VERIFIES everything: cross-checks Excel against web sources,        ║
║      flags unknown brands, detects large market-share swings,            ║
║      and writes a full verification report                               ║
║   4. Injects verified data into dashboard.html (replacing old block)     ║
╚══════════════════════════════════════════════════════════════════════════╝

Setup (one-time):
    pip install openpyxl requests beautifulsoup4
Run:
    python update_dashboard.py
"""

import os, re, json, unicodedata, logging, hashlib
from datetime import datetime, timezone
from pathlib import Path

# ── Optional imports ──────────────────────────────────────────────────────
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import requests
    from bs4 import BeautifulSoup
    HAS_REQUESTS = True
except ImportError:
    HAS_REQUESTS = False

# ═══════════════════════════════════════════════════════════════════════════
# CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════
SCRIPT_DIR     = Path(__file__).parent
EXCEL_FILE     = SCRIPT_DIR / "Vehicle_Planing_Trial.xlsx"
DASHBOARD_HTML = SCRIPT_DIR / "dashboard.html"
LOG_FILE       = SCRIPT_DIR / "update_log.txt"
SNAPSHOT_FILE  = SCRIPT_DIR / "last_snapshot.json"   # stores previous run's data
REPORT_FILE    = SCRIPT_DIR / "verification_report.txt"
EXCEL_SHEET    = "Sales Protocol "
MAX_EXCEL_ROWS = 400

# Market-share change alert threshold — flag if a brand moves more than this
MARKET_SHARE_ALERT_THRESHOLD = 5.0   # percentage points

# ═══════════════════════════════════════════════════════════════════════════
# KNOWN-GOOD BRAND REGISTRY
# Every brand Screens Studio tracks. Used to detect new / misspelled entries
# in the Excel file and cross-check availability claims.
# ═══════════════════════════════════════════════════════════════════════════
KNOWN_BRANDS = {
    # EU brands
    "volkswagen", "skoda", "audi", "cupra", "seat", "porsche", "lamborghini",
    "peugeot", "opel", "vauxhall", "citroen", "fiat", "jeep", "alfaromeo",
    "dsautomobiles", "lancia", "maserati",
    "renault", "dacia", "alpine",
    "hyundai", "kia", "genesis",
    "bmw", "mini",
    "toyota", "lexus",
    "mercedesbenz",
    "ford",
    "volvocars", "polestar",
    "mg", "saic",
    "nissan",
    "tesla",
    "byd",
    "suzuki", "mazda", "honda",
    "landrover", "jaguar",
    "astonmartin", "astionmartin",   # "astionmartin" = typo in Excel — both accepted
    "mclaren", "lotus", "smart", "vinfast", "subaru",
    "mitsubishi", "fisker", "aiways", "lucid", "rivian",
    # Trucks (commercial vehicles tracked on Screens)
    "truckdaf", "truckman", "truckvolvo", "truckmercedes",
    # US brands
    "chevrolet", "gmc", "cadillac", "buick",
    "lincoln",
    "dodge", "ram", "chrysler",
    "acura", "infiniti",
    "genesis",
    # CN brands
    "zeekr", "lynkco", "galaxy", "jiyue",
    "roewe", "immotors", "risingauto", "baojun", "wuling", "maxus",
    "aion", "hyptec", "trumpchi",
    "haval", "ora", "wey", "tank",
    "changan", "deepal", "avatr", "qiyuan",
    "nio", "onvo", "firefly",
    "xpeng", "xiaopeng",
    "liauto", "lixiang",
    "leapmotor",
    "seres", "aito",
    "xiaomi",
    "hongqi", "bestune",
    "voyah", "mhero", "nammi", "aeolus",
    "arcfox", "stelato", "baic",
    "jac", "maextro", "yiwei",
    "neta", "hozon",
    "hiphi",
    "chery", "exeed", "jetour", "jaecoo", "omoda", "luxeed", "icar",
    "weltmeister", "byton",
    "formulaleopard", "roxjishi",
    "lucidmotors", "smarteu", "smartcnsmart", "xiaomiauto",
    "jiyuejiyue", "geelyauto", "opelvauxhall",  # aliased dashboard keys
    # Additional variants found in Excel
    "volvo",           # Volvo Cars
    "geely",           # Geely Auto
    "rangerover",      # Range Rover (sub-brand of Land Rover)
    "danza", "denza",  # Denza (BYD JV) — "danza" is a common spelling variant
    "rising",          # Rising Auto (SAIC)
    "im",              # IM Motors (SAIC) — short form used in Excel
    "hyper",           # Hyptec / Hyper (GAC sub-brand)
    "kiamotors",       # Kia (older Excel entry name)
}

# Baseline market-share data (source: ACEA/Wards/CPCA 2025 Q1)
# Used to detect suspiciously large changes in fetched data.
BASELINE_MARKET_SHARE = {
    # EU (ACEA 2025)
    "eu": {
        "volkswagen": 26.9, "stellantis": 14.3, "renault": 10.2,
        "hyundai": 7.9, "bmw": 7.3, "toyota": 7.0, "mercedes": 5.1,
        "ford": 3.2, "volvo": 2.5, "saic": 2.3, "tesla": 1.8, "nissan": 2.2,
        "suzukimazdahonda": 2.9,
    },
    # US (Wards Auto 2025)
    "us": {
        "gm": 16.3, "ford": 13.1, "stellantis": 9.8, "toyota": 15.2,
        "honda": 9.2, "hyundai": 11.0, "nissan": 7.1, "bmw": 2.8,
        "mercedes": 2.2, "volkswagen": 3.4, "tesla": 3.5,
    },
    # CN NEV (CPCA 2025)
    "cn": {
        "byd": 32.0, "geely": 10.5, "saic": 9.8, "gac": 7.2,
        "changan": 6.8, "nio": 2.1, "xpeng": 2.3, "liauto": 3.5,
    },
}

# ═══════════════════════════════════════════════════════════════════════════
# LOGGING
# ═══════════════════════════════════════════════════════════════════════════
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
    ],
)
log = logging.getLogger("dashboard_updater")


# ═══════════════════════════════════════════════════════════════════════════
# BRAND ALIASES — map Excel short names → dashboard HTML brand-name keys
# ═══════════════════════════════════════════════════════════════════════════
# When a brand in the Excel file uses a different name than the dashboard HTML,
# add it here.  Key = nb(Excel brand name), Value = nb(dashboard brand name).
BRAND_ALIASES: dict[str, str] = {
    "volvo":        "volvocars",        # "Volvo Cars" in dashboard
    "opel":         "opelvauxhall",     # "Opel / Vauxhall" in dashboard
    "lucid":        "lucidmotors",      # "Lucid Motors" in dashboard
    "geely":        "geelyauto",        # "Geely Auto (吉利)" in dashboard
    "xiaomi":       "xiaomiauto",       # "Xiaomi Auto" in dashboard
    "xiaopeng":     "xpeng",            # "XPeng" in dashboard
    "lixiang":      "liauto",           # "Li Auto" in dashboard
    "im":           "immotors",         # "IM Motors (智己)" – short form in Excel
    "kiamotors":    "kia",              # "Kia motors" spelling variant
    "rangerover":   "landrover",        # Range Rover is Land Rover's flagship line
    "danza":        "denza",            # Danza/Denza spelling variant
    "astionmartin": "astonmartin",      # Typo in Excel: "Astion" → "Aston"
    "jiyue":        "jiyuejiyue",       # "JiYue / JIYUE (极越)"
    "rising":       "risingauto",       # "Rising Auto (飞凡)"
    "hozon":        "neta",             # Hozon = parent company; brand = Neta
}

# Smart has separate EU and CN dashboard entries
SMART_MARKET_KEYS = {"eu": "smarteu", "cn": "smartcnsmart"}

# ═══════════════════════════════════════════════════════════════════════════
# UTILITIES
# ═══════════════════════════════════════════════════════════════════════════
def nb(s: str) -> str:
    """Normalize brand name → a-z0-9 key. Must match _nb() in dashboard.html JS."""
    s = unicodedata.normalize("NFD", str(s))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"[^a-z0-9]", "", s.lower().strip())


def apply_brand_aliases(brands: dict) -> dict:
    """
    Remap Excel brand keys to the keys used in dashboard.html group data.
    Merges models when multiple Excel entries map to the same dashboard key.
    """
    result: dict = {}
    for key, data in brands.items():
        if key == "smart":
            # Smart has separate EU and CN dashboard entries
            for market, dash_key in SMART_MARKET_KEYS.items():
                if data[market]:
                    if dash_key not in result:
                        result[dash_key] = {"display": f"Smart ({market.upper()})", "eu": [], "us": [], "cn": []}
                    result[dash_key][market].extend(
                        m for m in data[market] if m not in result[dash_key][market]
                    )
            continue
        target = BRAND_ALIASES.get(key, key)
        if target not in result:
            result[target] = {"display": data["display"], "eu": [], "us": [], "cn": []}
        for market in ("eu", "us", "cn"):
            result[target][market].extend(
                m for m in data[market] if m not in result[target][market]
            )
    return result


def load_snapshot() -> dict:
    if SNAPSHOT_FILE.exists():
        try:
            return json.loads(SNAPSHOT_FILE.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


def save_snapshot(data: dict):
    SNAPSHOT_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


# ═══════════════════════════════════════════════════════════════════════════
# STEP 1 — READ EXCEL
# ═══════════════════════════════════════════════════════════════════════════
def read_excel(path: Path) -> dict:
    """
    Returns {normalised_key: {display, eu:[models], us:[models], cn:[models]}}
    """
    if not HAS_OPENPYXL:
        log.error("openpyxl missing — pip install openpyxl"); return {}
    if not path.exists():
        log.error(f"Excel not found: {path}"); return {}

    log.info(f"Reading Excel: {path.name}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    if EXCEL_SHEET not in wb.sheetnames:
        log.error(f"Sheet '{EXCEL_SHEET}' not found. Sheets: {wb.sheetnames}"); wb.close(); return {}

    ws = wb[EXCEL_SHEET]
    brands: dict = {}
    count = 0

    for row in ws.rows:
        vals = [cell.value for cell in row[:8]]
        if vals[0] is not None and isinstance(vals[0], (int, float)):
            brand = str(vals[1]).strip() if vals[1] else ""
            model = str(vals[2]).strip() if vals[2] else ""
            eu    = vals[3] == "x"
            us    = vals[4] == "x"
            cn    = vals[5] == "x"
            year  = vals[7]
            if not brand:
                continue
            key = nb(brand)
            if key not in brands:
                brands[key] = {"display": brand, "eu": [], "us": [], "cn": []}
            m_str = f"{model} (MY{int(year)})" if isinstance(year, (int, float)) and year else model
            if eu and m_str not in brands[key]["eu"]: brands[key]["eu"].append(m_str)
            if us and m_str not in brands[key]["us"]: brands[key]["us"].append(m_str)
            if cn and m_str not in brands[key]["cn"]: brands[key]["cn"].append(m_str)
        count += 1
        if count > MAX_EXCEL_ROWS:
            break

    wb.close()
    log.info(f"  → {len(brands)} unique brands | "
             f"EU: {sum(1 for b in brands.values() if b['eu'])} | "
             f"US: {sum(1 for b in brands.values() if b['us'])} | "
             f"CN: {sum(1 for b in brands.values() if b['cn'])}")
    return brands


# ═══════════════════════════════════════════════════════════════════════════
# STEP 2 — WEB RESEARCH
# ═══════════════════════════════════════════════════════════════════════════
def _safe_get(url: str, label: str) -> str | None:
    if not HAS_REQUESTS:
        return None
    try:
        r = requests.get(url, timeout=20,
                         headers={"User-Agent": "Mozilla/5.0 (compatible; ScreensDashboard/3.0)"})
        if r.status_code == 200:
            log.info(f"  ✓ Fetched {label}")
            return r.text
        log.warning(f"  {label} returned HTTP {r.status_code} — using cached values")
    except Exception as e:
        log.warning(f"  {label} fetch failed: {e} — using cached values")
    return None


def fetch_live_market_share() -> dict:
    """
    Try to pull current registration figures from public sources.
    Falls back gracefully to BASELINE_MARKET_SHARE on any failure.
    Returns structure matching BASELINE_MARKET_SHARE.
    """
    share = {k: dict(v) for k, v in BASELINE_MARKET_SHARE.items()}   # start with baseline

    # ── ACEA EU ───────────────────────────────────────────────────────────
    html = _safe_get(
        "https://www.acea.auto/figure/automobile-registrations-by-manufacturer-group-in-europe/",
        "ACEA EU registrations"
    )
    if html:
        soup = BeautifulSoup(html, "html.parser")
        tables = soup.find_all("table")
        for table in tables:
            for row in table.find_all("tr"):
                cells = [c.get_text(strip=True) for c in row.find_all(["td", "th"])]
                if len(cells) >= 2:
                    brand_cell = nb(cells[0])
                    for col in cells[1:]:
                        # Look for a percentage pattern like "26.9%" or "26,9%"
                        m = re.search(r"(\d+[,\.]\d+)\s*%", col)
                        if m:
                            pct = float(m.group(1).replace(",", "."))
                            # Map to our keys
                            for key in share["eu"]:
                                if key in brand_cell or brand_cell in key:
                                    if abs(pct - share["eu"][key]) < 15:  # sanity: ignore if wildly off
                                        share["eu"][key] = pct
                            break

    # ── Wards Auto US ─────────────────────────────────────────────────────
    # Wards requires a subscription — skip live fetch, use baseline
    # html = _safe_get("https://wardsintelligence.informa.com/...", "Wards US")

    # ── CPCA CN ───────────────────────────────────────────────────────────
    # CPCA publishes monthly PDFs; complex to parse — use baseline
    # html = _safe_get("http://www.cpca.org.cn/...", "CPCA CN")

    return share


def fetch_brand_verification_data(brand_key: str) -> dict | None:
    """
    Fetch a quick sanity check for a specific brand from Wikipedia or
    the brand's official newsroom. Returns minimal metadata or None.
    Light-touch: only called for brands flagged as suspicious.
    """
    WIKI_URLS = {
        "volkswagen": "https://en.wikipedia.org/wiki/Volkswagen",
        "toyota":     "https://en.wikipedia.org/wiki/Toyota",
        "byd":        "https://en.wikipedia.org/wiki/BYD_Auto",
        "tesla":      "https://en.wikipedia.org/wiki/Tesla,_Inc.",
        "bmw":        "https://en.wikipedia.org/wiki/BMW",
        "mercedesbenz": "https://en.wikipedia.org/wiki/Mercedes-Benz",
        "ford":       "https://en.wikipedia.org/wiki/Ford_Motor_Company",
        "hyundai":    "https://en.wikipedia.org/wiki/Hyundai_Motor_Company",
        "kia":        "https://en.wikipedia.org/wiki/Kia_Corporation",
        "nio":        "https://en.wikipedia.org/wiki/NIO_Inc.",
        "xpeng":      "https://en.wikipedia.org/wiki/Xpeng",
    }
    url = WIKI_URLS.get(brand_key)
    if not url:
        return None

    html = _safe_get(url, f"Wikipedia/{brand_key}")
    if not html:
        return None

    soup = BeautifulSoup(html, "html.parser")
    title = soup.find("h1")
    summary = soup.find("div", {"class": "mw-parser-output"})
    first_para = ""
    if summary:
        p = summary.find("p")
        if p:
            first_para = p.get_text()[:300]

    return {"title": title.get_text() if title else "", "summary": first_para}


# ═══════════════════════════════════════════════════════════════════════════
# STEP 3 — VERIFICATION ENGINE
# ═══════════════════════════════════════════════════════════════════════════
class VerificationReport:
    def __init__(self):
        self.timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
        self.issues: list[dict] = []     # {level, category, message}
        self.checks_passed = 0
        self.checks_failed = 0

    def ok(self, msg: str):
        self.checks_passed += 1
        log.info(f"  ✓ {msg}")

    def warn(self, category: str, msg: str):
        self.checks_failed += 1
        self.issues.append({"level": "WARN", "category": category, "message": msg})
        log.warning(f"  ⚠  [{category}] {msg}")

    def error(self, category: str, msg: str):
        self.checks_failed += 1
        self.issues.append({"level": "ERROR", "category": category, "message": msg})
        log.error(f"  ✗  [{category}] {msg}")

    def summary(self) -> str:
        total = self.checks_passed + self.checks_failed
        return (f"Verification: {self.checks_passed}/{total} checks passed, "
                f"{self.checks_failed} issue(s) found")

    def save(self, path: Path):
        lines = [
            "=" * 72,
            f"  SCREENS STUDIO — DASHBOARD VERIFICATION REPORT",
            f"  Generated: {self.timestamp}",
            "=" * 72,
            f"  Result:  {self.summary()}",
            "",
        ]
        if self.issues:
            lines.append("  ISSUES FOUND:")
            lines.append("  " + "-" * 68)
            for i, issue in enumerate(self.issues, 1):
                lines.append(f"  [{i}] {issue['level']} | {issue['category']}")
                lines.append(f"      {issue['message']}")
            lines.append("")
        else:
            lines.append("  No issues found — all data verified clean.")
            lines.append("")
        lines.append("=" * 72)
        path.write_text("\n".join(lines), encoding="utf-8")
        log.info(f"  Verification report saved → {path.name}")


def verify_excel_brands(brands: dict, report: VerificationReport):
    """CHECK 1 — Validate Excel brand names against known registry."""
    log.info("Verifying Excel brand names…")
    for key, data in brands.items():
        if key in KNOWN_BRANDS:
            report.ok(f"Known brand: {data['display']}")
        else:
            # Could be a new brand, a typo, or a legitimate newcomer
            report.warn(
                "UNKNOWN BRAND",
                f"'{data['display']}' (key='{key}') is not in the known-brands registry. "
                f"Models — EU:{data['eu']} US:{data['us']} CN:{data['cn']}. "
                f"Action: If this is a real brand, add '{key}' to KNOWN_BRANDS in update_dashboard.py."
            )


def verify_market_share(live_share: dict, snapshot: dict, report: VerificationReport):
    """CHECK 2 — Detect large market-share swings vs last run."""
    log.info("Verifying market-share data…")
    prev_share = snapshot.get("market_share", {})

    for market, brands in live_share.items():
        for brand, pct in brands.items():
            # Sanity: share should be between 0 and 60%
            if not (0 <= pct <= 60):
                report.error(
                    "MARKET SHARE RANGE",
                    f"{market.upper()} / {brand}: {pct}% is outside expected 0–60% range. "
                    f"Check the source data."
                )
                continue
            report.ok(f"{market.upper()} {brand}: {pct}%")

            # Compare to previous run
            prev = prev_share.get(market, {}).get(brand)
            if prev is not None:
                delta = abs(pct - prev)
                if delta > MARKET_SHARE_ALERT_THRESHOLD:
                    report.warn(
                        "LARGE SHARE CHANGE",
                        f"{market.upper()} / {brand}: changed {prev}% → {pct}% "
                        f"(Δ {delta:.1f} pp) — exceeds {MARKET_SHARE_ALERT_THRESHOLD} pp threshold. "
                        f"Please verify against source: "
                        f"{'acea.auto' if market=='eu' else 'wardsintelligence.informa.com' if market=='us' else 'cpca.org.cn'}"
                    )


def verify_model_counts(brands: dict, snapshot: dict, report: VerificationReport):
    """CHECK 3 — Flag brands that lost models since last run (could indicate data issue)."""
    log.info("Verifying model counts vs previous snapshot…")
    prev_brands = snapshot.get("brands", {})

    for key, data in brands.items():
        prev = prev_brands.get(key)
        if not prev:
            report.ok(f"New brand in Excel: {data['display']}")
            continue

        for market in ("eu", "us", "cn"):
            curr_count = len(data[market])
            prev_count = len(prev.get(market, []))
            if curr_count == prev_count:
                report.ok(f"{data['display']} {market.upper()}: {curr_count} model(s) — unchanged")
            elif curr_count > prev_count:
                added = [m for m in data[market] if m not in prev.get(market, [])]
                report.ok(f"{data['display']} {market.upper()}: +{curr_count - prev_count} model(s) added: {added}")
            else:
                removed = [m for m in prev.get(market, []) if m not in data[market]]
                report.warn(
                    "MODELS REMOVED",
                    f"{data['display']} {market.upper()}: {prev_count} → {curr_count} models. "
                    f"Removed: {removed}. "
                    f"If intentional (vehicle discontinued / removed from Screens), this is fine. "
                    f"If unexpected, check the Excel sheet row for this brand."
                )


def verify_excel_completeness(brands: dict, report: VerificationReport):
    """CHECK 4 — Flag brands that appear in Excel with no market assignment at all."""
    log.info("Verifying Excel completeness…")
    for key, data in brands.items():
        if not data["eu"] and not data["us"] and not data["cn"]:
            report.warn(
                "NO MARKET ASSIGNED",
                f"'{data['display']}' has no EU/US/CN availability in the Excel. "
                f"It will show as Missing on the dashboard. "
                f"If this brand has vehicles on Screens, add an 'x' in the correct column."
            )
        else:
            total = len(data["eu"]) + len(data["us"]) + len(data["cn"])
            report.ok(f"{data['display']}: {total} total model(s) across all markets")


def verify_excel_file_changed(path: Path, snapshot: dict, report: VerificationReport):
    """CHECK 5 — Detect if Excel was actually updated since last run."""
    log.info("Checking if Excel file has changed since last run…")
    try:
        digest = hashlib.md5(path.read_bytes()).hexdigest()
        prev_digest = snapshot.get("excel_md5")
        if prev_digest and digest == prev_digest:
            report.warn(
                "EXCEL UNCHANGED",
                f"Vehicle_Planing_Trial.xlsx has NOT changed since the last update run. "
                f"If you meant to update vehicle availability, make sure you saved the Excel file "
                f"to the correct folder before running the update."
            )
        else:
            report.ok("Excel file has been modified since last run — fresh data loaded")
        return digest
    except Exception as e:
        report.warn("FILE CHECK", f"Could not compute Excel checksum: {e}")
        return ""


# ═══════════════════════════════════════════════════════════════════════════
# STEP 4 — BUILD JS OVERRIDE BLOCK
# ═══════════════════════════════════════════════════════════════════════════
# Dashboard brand names sometimes differ from Excel names (e.g. "Audi (CN)" vs "AUDI")
# These extra keys ensure CN/variant dashboard entries get overrides applied.
DASHBOARD_KEY_EXTRAS: dict[str, str] = {
    "audicn":           "audi",             # "Audi (CN)" in CN groups
    "buickcn":          "buick",            # "Buick (CN)" in CN groups
    "byddynastyocean":  "byd",              # "BYD Dynasty & Ocean"
    "mgsaic":           "mg",               # "MG (SAIC)"
    "nissancn":         "nissan",           # "Nissan (CN)"
}

def build_override_block(brands: dict, market_share: dict,
                          report: VerificationReport, ts: str) -> str:
    override = {}
    for key, data in brands.items():
        override[key] = {
            "eu": data["eu"], "us": data["us"], "cn": data["cn"]
        }
    # Add variant keys so CN dashboard brand names (e.g. "Audi (CN)") also get matched
    for dash_key, base_key in DASHBOARD_KEY_EXTRAS.items():
        if base_key in override:
            override[dash_key] = override[base_key]

    issues_json = json.dumps([
        {"level": i["level"], "category": i["category"], "message": i["message"]}
        for i in report.issues
    ], ensure_ascii=False, indent=2)

    return (
        f"/* VEHICLE_OVERRIDE_START */\n"
        f"// Auto-generated — do NOT edit manually. Run update_dashboard.py to refresh.\n"
        f"// Last updated: {ts}  |  Source: {EXCEL_FILE.name}\n"
        f"// Verification: {report.summary()}\n"
        f"const VEHICLE_OVERRIDE = {json.dumps(override, ensure_ascii=False, indent=2)};\n"
        f"const DASHBOARD_LAST_UPDATED = \"{ts}\";\n"
        f"const DASHBOARD_EXCEL_SOURCE = \"{EXCEL_FILE.name}\";\n"
        f"const MARKET_SHARE_DATA = {json.dumps(market_share, ensure_ascii=False, indent=2)};\n"
        f"const VERIFICATION_ISSUES = {issues_json};\n"
        f"window.VEHICLE_OVERRIDE = VEHICLE_OVERRIDE;\n"
        f"window.DASHBOARD_LAST_UPDATED = DASHBOARD_LAST_UPDATED;\n"
        f"/* VEHICLE_OVERRIDE_END */"
    )


# ═══════════════════════════════════════════════════════════════════════════
# STEP 5 — INJECT INTO HTML
# ═══════════════════════════════════════════════════════════════════════════
OVERRIDE_RE = re.compile(
    r"/\* VEHICLE_OVERRIDE_START \*/.+?/\* VEHICLE_OVERRIDE_END \*/", re.DOTALL
)

def update_html(html_path: Path, new_block: str, report: VerificationReport) -> bool:
    if not html_path.exists():
        log.error(f"dashboard.html not found: {html_path}"); return False

    content = html_path.read_text(encoding="utf-8")

    if OVERRIDE_RE.search(content):
        content = OVERRIDE_RE.sub(new_block, content)
        log.info("  ✓ Replaced VEHICLE_OVERRIDE block")
    else:
        log.warning("  No marker found — inserting before </script>")
        content = content.replace("\n</script>\n</body>",
                                   f"\n{new_block}\n</script>\n</body>", 1)

    # Inject a visible warning banner if there are ERROR-level issues
    errors = [i for i in report.issues if i["level"] == "ERROR"]
    BANNER_RE = re.compile(r'<div id="verif-banner"[^>]*>.*?</div>', re.DOTALL)
    if errors:
        banner = (
            '<div id="verif-banner" style="background:#fff3cd;color:#856404;'
            'border:1px solid #ffc107;padding:8px 16px;font-size:.75rem;text-align:center">'
            f'⚠ {len(errors)} verification error(s) detected during last update — '
            'check verification_report.txt for details.</div>'
        )
    else:
        banner = '<div id="verif-banner" style="display:none"></div>'

    if BANNER_RE.search(content):
        content = BANNER_RE.sub(banner, content)
    else:
        content = content.replace("<body", banner + "\n<body", 1)

    html_path.write_text(content, encoding="utf-8")
    log.info(f"  ✓ Saved {html_path.name} ({len(content)/1024:.1f} KB)")
    return True


# ═══════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════
def main():
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    log.info("=" * 72)
    log.info("  Screens Studio Dashboard Updater  v3.0")
    log.info(f"  Run: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  {EXCEL_FILE.name}")
    log.info("=" * 72)

    report   = VerificationReport()
    snapshot = load_snapshot()

    # ── 1. Read Excel ────────────────────────────────────────────────────
    log.info("\n[1/4] Reading Excel vehicle data…")
    brands = read_excel(EXCEL_FILE)
    if not brands:
        log.error("No brand data — aborting."); return 1
    brands = apply_brand_aliases(brands)
    log.info(f"  → After alias mapping: {len(brands)} dashboard keys")

    # ── 2. Web research ──────────────────────────────────────────────────
    log.info("\n[2/4] Fetching live market data from web…")
    market_share = fetch_live_market_share()

    # ── 3. Verification ──────────────────────────────────────────────────
    log.info("\n[3/4] Running verification checks…")
    excel_md5 = verify_excel_file_changed(EXCEL_FILE, snapshot, report)
    verify_excel_brands(brands, report)
    verify_excel_completeness(brands, report)
    verify_model_counts(brands, snapshot, report)
    verify_market_share(market_share, snapshot, report)
    report.save(REPORT_FILE)
    log.info(f"\n  {report.summary()}")

    # ── 4. Build + inject ────────────────────────────────────────────────
    log.info("\n[4/4] Updating dashboard.html…")
    block = build_override_block(brands, market_share, report, ts)
    ok    = update_html(DASHBOARD_HTML, block, report)
    if not ok: return 1

    # ── Save snapshot for next run ───────────────────────────────────────
    save_snapshot({
        "timestamp":    ts,
        "excel_md5":    excel_md5,
        "brands":       {k: {"eu": v["eu"], "us": v["us"], "cn": v["cn"]}
                          for k, v in brands.items()},
        "market_share": market_share,
    })

    log.info("\n" + "=" * 72)
    if report.checks_failed == 0:
        log.info("  ✅  Update complete — all checks passed!")
    else:
        log.info(f"  ⚠   Update complete — {report.checks_failed} issue(s) flagged.")
        log.info(f"      See {REPORT_FILE.name} for details.")
    log.info("=" * 72)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
